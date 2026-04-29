#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
license_lite.py - 启动Lite授权工具并自动完成机器标识、授权期限输入，生成授权码。

用法：
    python license_lite.py [--root] [--xlsx 路径] [--xlsx-start-row N] [MACHINE_ID]

选项：
    --root              以 root 模式启动，启动参数从 MYSHELL/config/private/password.json 的 lite-forever 读取
    --xlsx PATH         多个授权：从 xlsx 读取 D 列机器标识，生成后写入 E 列授权码（与交互选「多个授权」相同）
    --xlsx-start-row N  数据起始行号，默认 5（Excel 行号，从 1 起算）
    MACHINE_ID          单个授权时直接指定机器标识；与 --xlsx 互斥

交互：
    未传 MACHINE_ID 且未传 --xlsx：先选 1 单个 / 2 多个（questionary 菜单）→ 多个时 Windows 优先弹出系统文件框选 xlsx；
    机器标识、起始行、路径回退均用标准 input（避免 questionary 文本框无法粘贴）。
    授权期限：直接回车则使用「今天 + 90 天」的 YYYY-MM-DD。
    如果目标窗口已存在，则直接使用现有窗口，不再重复启动程序。
"""

import os
import sys
import time
import argparse
import subprocess
import re
import json
from datetime import datetime, timedelta
from pathlib import Path

import win32gui
import uiautomation as auto
import pyperclip

# ==================== 配置区域（请根据UI特征修改） ====================
# 窗口标题
WINDOW_TITLE = "K-RPA Lite授权"

# 授权密码（填入密码框的字符串）
PASSWORD = "kingautomate"

# 等待窗口出现的最长时间（秒）
MAX_WAIT_SECONDS = 15

# ----- 控件定位特征（修改这些以匹配实际UI） -----
# 机器标识框
MACHINE_CLASS = "TEdit"
MACHINE_IS_PASSWORD = False
MACHINE_IS_READONLY = False
MACHINE_HELP_TEXT = ""                 # 帮助文本为空

# 授权期限框
EXPIRY_CLASS = "TEdit"
EXPIRY_HELP_TEXT_KEYWORD = "正确日期格式"   # 帮助文本包含的关键词

# 授权密码框
PASSWORD_CLASS = "TEdit"
PASSWORD_IS_PASSWORD = True

# 授权码结果框
RESULT_CLASS = "TMemo"

# 生成按钮
GENERATE_BUTTON_NAME = "生成"

# ----- 备用定位（当上述条件匹配失败时使用）-----
# 是否启用备用方案（基于索引或关键词）
ENABLE_FALLBACK = True
# 备用方案中假设的索引（例如：授权期限框可能是第一个TEdit）
FALLBACK_EXPIRY_INDEX = 0
# 备用方案中根据“日期”关键词搜索
FALLBACK_EXPIRY_KEYWORD = "日期"

# ----- 可执行文件路径 -----
DEFAULT_EXE_RELATIVE = Path("KingAutomate/Licenses/LiteLicense/LiteLicense.exe")
PROJECT_ROOT_ENV_VAR = "AOM_ROOT"          # 可选环境变量，用于覆盖项目根目录
DEFAULT_PROJECT_ROOT = r"D:\Code\aom"      # 默认项目根目录
# ============================================================

def resolve_password_file():
    """从环境变量 MYSHELL 获取密码文件路径：$MYSHELL/config/private/password.json"""
    myshell = os.environ.get('MYSHELL')
    if not myshell:
        print("错误: 环境变量 MYSHELL 未设置，无法定位密码文件。", file=sys.stderr)
        return None
    password_file = Path(myshell) / "config/private/password.json"
    return password_file

def get_root_startup_param(password_file):
    """从密码文件中读取 lite-forever.password 作为 root 模式启动参数"""
    if not password_file.exists():
        print(f"错误: 密码文件不存在: {password_file}", file=sys.stderr)
        return None
    try:
        with open(password_file, 'r', encoding='utf-8-sig') as f:
            config = json.load(f)
    except json.JSONDecodeError as e:
        print(f"错误: 密码文件格式错误，不是有效的 JSON: {e}", file=sys.stderr)
        return None
    except Exception as e:
        print(f"错误: 读取密码文件失败: {e}", file=sys.stderr)
        return None

    lite_forever = config.get('lite-forever')
    if not lite_forever or not isinstance(lite_forever, dict):
        print("错误: 密码文件中未找到 'lite-forever' 对象。", file=sys.stderr)
        return None
    param = lite_forever.get('password')
    if not param or not isinstance(param, str):
        print("错误: 'lite-forever' 中缺少有效的 'password' 字段。", file=sys.stderr)
        return None
    return param

def find_window_handle(title):
    """根据窗口标题查找窗口句柄（精确匹配）"""
    def enum_callback(hwnd, hwnds):
        if win32gui.IsWindowVisible(hwnd) and win32gui.GetWindowText(hwnd) == title:
            hwnds.append(hwnd)
        return True
    hwnds = []
    win32gui.EnumWindows(enum_callback, hwnds)
    return hwnds[0] if hwnds else None

def get_edit_boxes_info(window):
    """获取窗口中所有编辑框的属性列表"""
    all_controls = window.GetChildren()
    edit_boxes = []
    for ctrl in all_controls:
        if ctrl.ControlType == auto.ControlType.EditControl:
            try:
                is_password = ctrl.GetPropertyValue(auto.PropertyId.IsPasswordProperty)
            except:
                is_password = 'N/A'
            try:
                class_name = ctrl.GetPropertyValue(auto.PropertyId.ClassNameProperty)
            except:
                class_name = 'N/A'
            automation_id = ctrl.AutomationId
            bounding_rect = ctrl.BoundingRectangle

            is_readonly = 'N/A'
            try:
                value_pattern = ctrl.GetValuePattern()
                is_readonly = value_pattern.IsReadOnly
            except:
                try:
                    is_readonly = ctrl.GetPropertyValue(auto.PropertyId.IsReadOnlyProperty)
                except:
                    pass

            help_text = ''
            try:
                help_text = ctrl.GetPropertyValue(auto.PropertyId.HelpTextProperty)
            except:
                pass

            edit_boxes.append({
                'control': ctrl,
                'index': len(edit_boxes),
                'is_password': is_password,
                'class_name': class_name,
                'automation_id': automation_id,
                'rect': bounding_rect,
                'is_readonly': is_readonly,
                'help_text': help_text
            })
    return edit_boxes

def input_to_control(ctrl, text, description):
    """向控件输入文本，尝试多种方法"""
    print(f"向 {description} 输入文本...")
    ctrl.SetFocus()
    time.sleep(0.2)

    # 尝试 ValuePattern.SetValue
    try:
        value_pattern = ctrl.GetValuePattern()
        value_pattern.SetValue(text)
        print(f"  ✅ 通过 ValuePattern 成功输入")
        return True
    except Exception as e:
        print(f"  ValuePattern 输入失败: {e}")

    # 尝试 SendKeys
    try:
        ctrl.SendKeys(text)
        print(f"  ✅ 通过 SendKeys 成功输入")
        return True
    except Exception as e:
        print(f"  SendKeys 输入失败: {e}")

    # 尝试 Click + SendKeys
    try:
        ctrl.Click()
        time.sleep(0.2)
        ctrl.SendKeys(text)
        print(f"  ✅ 通过 Click+SendKeys 成功输入")
        return True
    except Exception as e:
        print(f"  Click+SendKeys 输入失败: {e}")

    print(f"  ❌ 所有输入方法均失败")
    return False

def click_generate_button(window):
    """点击'生成'按钮，尝试多种方法"""
    print("\n正在查找'生成'按钮...")
    button = None

    # 方法1：通过 Name 定位
    try:
        button = window.ButtonControl(Name=GENERATE_BUTTON_NAME)
        if not button.Exists():
            button = None
    except:
        pass

    if not button:
        # 方法2：遍历查找 Name='生成' 的按钮
        try:
            for ctrl in window.GetChildren():
                if ctrl.ControlType == auto.ControlType.ButtonControl and ctrl.Name == GENERATE_BUTTON_NAME:
                    button = ctrl
                    break
        except:
            pass

    if not button:
        print(f"❌ 未能找到'{GENERATE_BUTTON_NAME}'按钮")
        return False

    print(f"✅ 找到'{GENERATE_BUTTON_NAME}'按钮，尝试点击...")
    button.SetFocus()
    time.sleep(0.3)

    # 点击尝试顺序
    try:
        button.Invoke()
        print("   ✅ 通过 InvokePattern 点击成功")
        return True
    except Exception as e:
        print(f"   InvokePattern 失败: {e}")

    try:
        legacy = button.GetLegacyIAccessiblePattern()
        if legacy:
            legacy.DoDefaultAction()
            print("   ✅ 通过 LegacyIAccessible.DoDefaultAction 点击成功")
            return True
    except Exception as e:
        print(f"   LegacyIAccessible 失败: {e}")

    try:
        rect = button.BoundingRectangle
        if rect:
            x = (rect.left + rect.right) // 2
            y = (rect.top + rect.bottom) // 2
            auto.Click(x, y)
            print(f"   ✅ 通过模拟鼠标点击 ({x}, {y}) 成功")
            return True
    except Exception as e:
        print(f"   模拟鼠标点击失败: {e}")

    try:
        button.SendKeys('{Space}')
        print("   ✅ 通过发送空格键点击成功")
        return True
    except Exception as e:
        print(f"   发送空格键失败: {e}")

    print("❌ 所有点击方法均失败")
    return False

def extract_license_code_from_window(window):
    """
    从授权码结果框（TMemo）读取文本，复制到剪贴板。
    返回授权码字符串；失败返回 None。
    """
    print("\n正在获取授权码结果...")
    all_controls = window.GetChildren()
    memo_ctrl = None

    for ctrl in all_controls:
        if ctrl.ControlType == auto.ControlType.EditControl:
            try:
                class_name = ctrl.GetPropertyValue(auto.PropertyId.ClassNameProperty)
                if class_name == RESULT_CLASS:
                    memo_ctrl = ctrl
                    break
            except Exception:
                continue

    if not memo_ctrl:
        print(f"❌ 未找到授权码结果框（ClassName={RESULT_CLASS}）")
        return None

    try:
        value_pattern = memo_ctrl.GetValuePattern()
        code_text = value_pattern.Value
        if code_text:
            code_text = str(code_text).strip()
            pyperclip.copy(code_text)
            print(f"✅ 已复制授权码到剪贴板，内容长度: {len(code_text)} 字符")
            print(f"   前50字符: {code_text[:50]}{'...' if len(code_text) > 50 else ''}")
            return code_text
        print("⚠️ 授权码内容为空，可能尚未生成")
        return None
    except Exception as e:
        print(f"获取授权码失败: {e}")
        return None


def copy_authorization_code_to_clipboard(window):
    """兼容旧调用：从窗口取授权码并复制到剪贴板，成功返回 True。"""
    return extract_license_code_from_window(window) is not None

def get_default_exe_path():
    """获取默认的 LiteLicense.exe 路径（支持环境变量覆盖）"""
    project_root = os.environ.get(PROJECT_ROOT_ENV_VAR, DEFAULT_PROJECT_ROOT)
    exe_path = Path(project_root) / DEFAULT_EXE_RELATIVE
    return exe_path

def wait_for_window(title, max_seconds=MAX_WAIT_SECONDS):
    """等待窗口出现，返回窗口句柄或None"""
    print(f"等待窗口 '{title}' 出现...")
    for i in range(int(max_seconds * 2)):  # 每0.5秒检查一次
        hwnd = find_window_handle(title)
        if hwnd:
            print(f"✅ 窗口已出现 (句柄: {hwnd})")
            return hwnd
        time.sleep(0.5)
    print(f"❌ 等待窗口超时 ({max_seconds}秒)")
    return None

def validate_date(date_str):
    """验证日期格式是否为 YYYY-MM-DD 且为有效日期"""
    if not re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
        return False
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
        return True
    except ValueError:
        return False


def default_expiry_date_days_from_today(days=90):
    """返回「今天 + days」的日期字符串 YYYY-MM-DD。"""
    d = datetime.now().date() + timedelta(days=days)
    return d.strftime("%Y-%m-%d")


def _questionary_available():
    try:
        import questionary  # noqa: F401
        return True
    except ImportError:
        return False


def choose_authorization_mode_interactive():
    """
    返回 'single' | 'multi' | None（取消）。
    优先使用 questionary（与 _tool 菜单一致）；否则回退到 input。
    """
    if _questionary_available():
        import questionary
        sel = questionary.select(
            "请选择授权方式：",
            choices=[
                "1. 单个授权",
                "2. 多个授权",
            ],
            instruction="(↑/↓ 选择，回车确认，Esc 取消)",
        ).ask()
        if sel is None:
            return None
        if sel.startswith("1"):
            return "single"
        if sel.startswith("2"):
            return "multi"
        return None

    print("请选择授权方式：")
    print("  1. 单个授权")
    print("  2. 多个授权")
    while True:
        choice = input("请输入 1 或 2: ").strip()
        if choice == "1":
            return "single"
        if choice == "2":
            return "multi"
        print("无效输入，请重新输入 1 或 2。")


def prompt_machine_id_single_interactive():
    """单个授权：读取一行机器标识（使用标准 input，避免 questionary 文本框无法粘贴）。"""
    print("请输入机器标识（可从客户处获得，支持粘贴后回车）：")
    return input("> ").strip()


def prompt_expiry_date_interactive(default_days=90):
    default_s = default_expiry_date_days_from_today(default_days)
    print(
        f"请输入授权期限（格式 YYYY-MM-DD，直接回车默认：今天起第 {default_days} 天 = {default_s}）："
    )
    while True:
        raw = input("> ").strip()
        expiry_date = raw if raw else default_s
        if validate_date(expiry_date):
            return expiry_date
        print("日期格式错误或无效，请重新输入（例如 2025-12-31）：")


def run_lite_authorization(machine_id, expiry_date, use_root):
    """
    对当前（或新启动的）授权窗口完成一次：填机器标识、期限、密码，生成并复制授权码。
    返回授权码文本；失败时进程内已 print，返回 None（部分路径会 sys.exit）。
    """
    hwnd = find_window_handle(WINDOW_TITLE)
    if hwnd:
        print(f"✅ 已存在窗口 '{WINDOW_TITLE}'，直接使用现有窗口。")
    else:
        exe_path = get_default_exe_path()
        if not exe_path.exists():
            print(f"错误: 程序文件不存在: {exe_path}", file=sys.stderr)
            sys.exit(1)

        if use_root:
            password_file = resolve_password_file()
            if password_file is None or not password_file.exists():
                print("错误: 密码文件不存在或环境变量 MYSHELL 未设置。", file=sys.stderr)
                sys.exit(1)
            startup_param = get_root_startup_param(password_file)
            if startup_param is None:
                print("无法获取 root 启动参数，退出。", file=sys.stderr)
                sys.exit(1)
            print(f"启动 LiteLicense.exe (Root模式)，参数: {startup_param}")
            subprocess.Popen([str(exe_path), startup_param])
        else:
            print("启动 LiteLicense.exe...")
            subprocess.Popen([str(exe_path)])

        hwnd = wait_for_window(WINDOW_TITLE)
        if not hwnd:
            print("无法找到授权窗口，退出。", file=sys.stderr)
            sys.exit(1)

    window = auto.ControlFromHandle(hwnd)
    window.SetFocus()
    time.sleep(0.5)

    edit_boxes = get_edit_boxes_info(window)
    print("\n当前窗口中所有编辑框：")
    for eb in edit_boxes:
        print(f"  索引 {eb['index']}: ClassName={eb['class_name']}, IsPassword={eb['is_password']}, IsReadOnly={eb['is_readonly']}, HelpText='{eb['help_text']}', AutomationId={eb['automation_id']}")

    machine_ctrl = None
    for eb in edit_boxes:
        if (eb['class_name'] == MACHINE_CLASS
                and eb['is_password'] is MACHINE_IS_PASSWORD
                and eb['is_readonly'] is MACHINE_IS_READONLY
                and (eb['help_text'] is None or eb['help_text'] == MACHINE_HELP_TEXT)):
            machine_ctrl = eb['control']
            print(f"\n✅ 找到机器标识框，索引 {eb['index']}")
            break

    expiry_ctrl = None
    for eb in edit_boxes:
        help_txt = eb.get('help_text') or ''
        if eb['class_name'] == EXPIRY_CLASS and EXPIRY_HELP_TEXT_KEYWORD in str(help_txt):
            expiry_ctrl = eb['control']
            print(f"✅ 找到授权期限框，索引 {eb['index']} (HelpText: {eb['help_text']})")
            break

    password_ctrl = None
    for eb in edit_boxes:
        if eb['class_name'] == PASSWORD_CLASS and eb['is_password'] is PASSWORD_IS_PASSWORD:
            password_ctrl = eb['control']
            print(f"✅ 找到授权密码框，索引 {eb['index']}")
            break

    if ENABLE_FALLBACK:
        if not machine_ctrl:
            candidates = [eb for eb in edit_boxes if eb['class_name'] == MACHINE_CLASS and eb['is_readonly'] is False and eb['is_password'] is False]
            if candidates:
                machine_ctrl = candidates[0]['control']
                print(f"\n⚠️ 使用备用方案，选择第一个可编辑{MACHINE_CLASS}作为机器标识框，索引 {candidates[0]['index']}")
        if not expiry_ctrl:
            candidates = [
                eb for eb in edit_boxes
                if eb['class_name'] == EXPIRY_CLASS and FALLBACK_EXPIRY_KEYWORD in str(eb.get('help_text') or '')
            ]
            if candidates:
                expiry_ctrl = candidates[0]['control']
                print(f"⚠️ 使用备用方案，根据'{FALLBACK_EXPIRY_KEYWORD}'关键词找到授权期限框，索引 {candidates[0]['index']}")
            else:
                if len(edit_boxes) > FALLBACK_EXPIRY_INDEX and edit_boxes[FALLBACK_EXPIRY_INDEX]['class_name'] == EXPIRY_CLASS:
                    expiry_ctrl = edit_boxes[FALLBACK_EXPIRY_INDEX]['control']
                    print(f"⚠️ 使用备用方案，假设索引{FALLBACK_EXPIRY_INDEX}为授权期限框")
        if not password_ctrl:
            candidates = [eb for eb in edit_boxes if eb['class_name'] == PASSWORD_CLASS and eb['is_password'] is True]
            if candidates:
                password_ctrl = candidates[0]['control']
                print(f"⚠️ 使用备用方案，选择第一个密码{PASSWORD_CLASS}作为授权密码框，索引 {candidates[0]['index']}")

    if not machine_ctrl:
        print("❌ 无法定位机器标识框")
        sys.exit(1)
    if not expiry_ctrl:
        print("❌ 无法定位授权期限框")
        sys.exit(1)
    if not password_ctrl:
        print("❌ 无法定位授权密码框")
        sys.exit(1)

    input_to_control(machine_ctrl, machine_id, "机器标识框")
    input_to_control(expiry_ctrl, expiry_date, "授权期限框")
    input_to_control(password_ctrl, PASSWORD, "授权密码框")

    click_success = click_generate_button(window)
    if not click_success:
        print("点击生成按钮失败，可能无法生成授权码。")

    print("\n等待授权码生成...")
    time.sleep(2)

    code = extract_license_code_from_window(window)
    print("\n✅ 本台机器操作完成。")
    return code


def _require_openpyxl():
    try:
        from openpyxl import load_workbook  # noqa: F401
        return True
    except ImportError:
        print("多个授权（xlsx）需要安装 openpyxl，请执行: pip install openpyxl", file=sys.stderr)
        return False


def load_xlsx_jobs(xlsx_path, start_row):
    """
    打开 xlsx，从第 start_row 行起读 D 列（机器标识），返回 (wb, ws, path, jobs)。
    jobs: [(excel_row_1based, machine_id), ...]，跳过 D 为空的行。
    """
    if not _require_openpyxl():
        sys.exit(1)
    from openpyxl import load_workbook

    path = Path(xlsx_path).expanduser().resolve()
    if not path.is_file():
        print(f"错误: 文件不存在: {path}", file=sys.stderr)
        sys.exit(1)
    if path.suffix.lower() not in (".xlsx", ".xlsm"):
        print("错误: 请使用 .xlsx 或 .xlsm 文件。", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filename=str(path), read_only=False, data_only=False)
    ws = wb.active
    col_machine = 4   # D
    jobs = []
    start = max(1, int(start_row))
    for row in range(start, ws.max_row + 1):
        raw = ws.cell(row=row, column=col_machine).value
        if raw is None:
            continue
        text = str(raw).strip()
        if not text:
            continue
        jobs.append((row, text))
    return wb, ws, path, jobs


def pick_xlsx_path_via_file_dialog():
    """图形界面选择 xlsx；取消或失败返回 None。"""
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        return None

    root = tk.Tk()
    root.withdraw()
    root.wm_attributes("-topmost", 1)
    try:
        path = filedialog.askopenfilename(
            parent=root,
            title="选择 xlsx（D 列=机器标识，E 列将写入授权码）",
            filetypes=[
                ("Excel 工作簿", "*.xlsx"),
                ("Excel 宏", "*.xlsm"),
                ("所有文件", "*.*"),
            ],
        )
    finally:
        root.destroy()
    return path.strip() if path else None


def prompt_xlsx_path_interactive():
    """
    多个授权：获取本地 xlsx 路径。
    Windows 下优先弹出系统文件选择框（无需在终端粘贴）；取消或失败则用 input。
    """
    print()
    print("提示：若不想交互选文件，可直接运行：")
    print('  _tool license-lite --xlsx "D:\\\\path\\\\file.xlsx"')
    print()

    if sys.platform == "win32":
        print("正在打开「选择 xlsx」窗口…（若取消，请在下面一行粘贴路径）")
        try:
            picked = pick_xlsx_path_via_file_dialog()
            if picked:
                print(f"已选择: {picked}")
                return picked
        except Exception as exc:
            print(f"(打开文件对话框失败: {exc})")

    print("请输入 .xlsx 的完整路径（在终端里可直接 Ctrl+V 粘贴）：")
    return input("> ").strip().strip('"')


def prompt_xlsx_start_row_interactive(default=5):
    """起始行号，默认 5（标准 input，避免 questionary 文本框问题）。"""
    raw = input(f"数据起始行号（Excel 行号，直接回车使用 {default}）：").strip()
    if not raw:
        return default
    try:
        n = int(raw)
        return max(1, n)
    except ValueError:
        print(f"无效数字，使用默认起始行 {default}。")
        return default


def main():
    parser = argparse.ArgumentParser(description="自动完成Lite授权工具的机器标识、授权期限输入和授权码生成")
    parser.add_argument("--root", action="store_true", help="以 root 模式启动（启动参数从配置文件读取）")
    parser.add_argument(
        "--xlsx",
        default=None,
        help="多个授权：xlsx 路径；D 列机器标识，E 列写入授权码（与交互选「多个授权」一致）",
    )
    parser.add_argument(
        "--xlsx-start-row",
        type=int,
        default=5,
        metavar="N",
        help="xlsx 数据起始行（Excel 行号，默认 5）",
    )
    parser.add_argument(
        "machine_id",
        nargs="?",
        default=None,
        help="单个授权时的机器标识；与 --xlsx 互斥",
    )
    args = parser.parse_args()
    use_root = args.root
    xlsx_arg = (args.xlsx or "").strip()
    machine_id_arg = (args.machine_id or "").strip()

    if xlsx_arg and machine_id_arg:
        print("错误: 不能同时指定 MACHINE_ID 与 --xlsx。", file=sys.stderr)
        sys.exit(2)

    xlsx_bundle = None  # (wb, ws, path, jobs)
    single_ids = None

    if xlsx_arg:
        wb, ws, path, jobs = load_xlsx_jobs(xlsx_arg, args.xlsx_start_row)
        if not jobs:
            print("错误: 从起始行起未在 D 列读到任何机器标识。", file=sys.stderr)
            sys.exit(1)
        xlsx_bundle = (wb, ws, path, jobs)
    elif machine_id_arg:
        single_ids = [machine_id_arg]
    else:
        mode = choose_authorization_mode_interactive()
        if mode is None:
            print("已取消。")
            sys.exit(0)
        if mode == "single":
            mid = prompt_machine_id_single_interactive()
            if not mid:
                print("错误: 机器标识不能为空。", file=sys.stderr)
                sys.exit(2)
            single_ids = [mid]
        else:
            xlsx_path = prompt_xlsx_path_interactive()
            if not xlsx_path:
                print("错误: xlsx 路径不能为空。", file=sys.stderr)
                sys.exit(2)
            start_row = prompt_xlsx_start_row_interactive(args.xlsx_start_row)
            wb, ws, path, jobs = load_xlsx_jobs(xlsx_path, start_row)
            if not jobs:
                print("错误: 从起始行起未在 D 列读到任何机器标识。", file=sys.stderr)
                sys.exit(1)
            xlsx_bundle = (wb, ws, path, jobs)

    expiry_date = prompt_expiry_date_interactive()

    if xlsx_bundle:
        wb, ws, path, jobs = xlsx_bundle
        total = len(jobs)
        col_license = 5  # E
        for idx, (row, mid) in enumerate(jobs):
            print(f"\n{'=' * 50}\n第 {idx + 1}/{total} 台 — Excel 第 {row} 行，机器标识: {mid}\n{'=' * 50}")
            code = run_lite_authorization(mid, expiry_date, use_root)
            if not code:
                print(f"❌ 第 {row} 行未获取到授权码，已中止。", file=sys.stderr)
                try:
                    wb.save(str(path))
                    print(f"已尽力保存当前进度到: {path}")
                except Exception as e:
                    print(f"保存 xlsx 失败: {e}", file=sys.stderr)
                sys.exit(1)
            ws.cell(row=row, column=col_license).value = code
            try:
                wb.save(str(path))
            except Exception as e:
                print(f"❌ 写入 Excel 失败: {e}", file=sys.stderr)
                sys.exit(1)
            print(f"✅ 已写入第 {row} 行 E 列并保存文件。")
        print(f"\n✅ 全部 {total} 行已处理: {path}")
    else:
        total = len(single_ids)
        for idx, mid in enumerate(single_ids):
            if total > 1:
                print(f"\n{'=' * 50}\n第 {idx + 1}/{total} 台 — 机器标识: {mid}\n{'=' * 50}")
            run_lite_authorization(mid, expiry_date, use_root)

        if total > 1:
            print(f"\n✅ 全部 {total} 台机器处理完毕。")
        else:
            print("\n✅ 所有操作完成。")


if __name__ == "__main__":
    main()